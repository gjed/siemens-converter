"""Tests for XLSX writer."""

from __future__ import annotations

import openpyxl
import pytest
from pathlib import Path

from siemens_converter.writer import write_xlsx
from siemens_converter.models import (
    ReportHeader,
    CentralMeter,
    WaterMeter,
    HeatAllocator,
    ParsedReport,
)


def _make_report():
    """Build a ParsedReport with 2 apartments for testing."""
    header = ReportHeader(
        "FC_report.xls",
        "2026-03-14",
        "16:45:03",
        "-",
        "EV123",
        "3.93",
        12,
    )
    central = [
        CentralMeter("Riscaldamento", 29213, 71714666, "2026/03/02"),
        CentralMeter("Sanitario", 31310, 71731475, "2026/03/02"),
    ]
    water = [
        WaterMeter("App, 01 Rossi", 1, 31.613, 23005747, "2026/03/02"),
        WaterMeter("App, 02 Bianchi", 2, 39.18, 22079950, "2026/03/02"),
    ]
    heat = [
        HeatAllocator("App, 01 Rossi", 1, 5.243, 45.72, 71714345, "2026/03/02"),
        HeatAllocator("App, 02 Bianchi", 2, 6.26, 59.69, 71714369, "2026/03/02"),
    ]
    return ParsedReport(header, central, water, heat)


def test_output_file_created(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    assert out.exists()


def test_central_meter_riscaldamento(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb.worksheets[0]
    assert ws["C28"].value == 29213


def test_central_meter_sanitario(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb.worksheets[0]
    assert ws["C31"].value == 31310


def test_reading_date(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb.worksheets[0]
    assert ws["D28"].value is not None


def test_heat_allocator_values(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb.worksheets[0]
    assert ws["C78"].value == 5243  # 5.243 MWh * 1000
    assert ws["C80"].value == 6260  # 6.26 MWh * 1000


def test_water_meter_values(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb.worksheets[0]
    assert ws["C130"].value == 31.613
    assert ws["C132"].value == 39.18


def test_afs_values(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb.worksheets[0]
    assert ws["C182"].value == 45.72
    assert ws["C184"].value == 59.69


def test_four_sheets(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 4


def test_dati_report_sheet_exists(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    assert "Dati Report" in wb.sheetnames


def test_dati_report_metadata(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Dati Report"]
    # Row 1: metadata labels in FC_report column positions
    assert ws["A1"].value == "Nome File"
    assert ws["B1"].value == "Data Report"
    assert ws["E1"].value == "Versione firmware"
    assert ws["H1"].value == "Numero di serie"
    # Row 1: group headers over data columns
    assert ws.cell(row=1, column=16).value == "Riparto"  # col P
    assert ws.cell(row=1, column=25).value == "Riparto"  # col Y
    assert ws.cell(row=1, column=27).value == "Conti separati a parte"  # col AA
    # Row 2: metadata values
    assert ws["A2"].value == "FC_report.xls"
    assert ws["H2"].value == "EV123"
    # Row 3: annotation sub-labels
    assert ws.cell(row=3, column=16).value == "energia termica"
    assert ws.cell(row=3, column=25).value == "volume acqua sanitaria"
    assert ws.cell(row=3, column=27).value == "volume acqua fredda"


def test_millesimali_names_from_report(tmp_path):
    out = tmp_path / "output.xlsx"
    write_xlsx(_make_report(), out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Tabelle millesimali"]
    assert ws["A4"].value == "App, 01 Rossi"
    assert ws["A5"].value == "App, 02 Bianchi"
    assert ws["B4"].value is None
    assert ws["C4"].value is None
    assert ws["E4"].value is None


def test_dati_report_full_fc_data(tmp_path):
    """Parse fixture -> verify Dati Report has all 38 columns from FC_report."""
    from siemens_converter.parser import parse_fc_report

    fixture = Path(__file__).parent / "fixtures" / "FC_report_TEST_9999_2026-01-15.xls"
    report = parse_fc_report(fixture)
    out = tmp_path / "output.xlsx"
    write_xlsx(report, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Dati Report"]

    # Row 4 should have all FC_report column headers
    assert ws.cell(row=4, column=1).value == "count"
    assert ws.cell(row=4, column=5).value == "device_description"  # col E
    assert ws.cell(row=4, column=16).value == "heat_energy"  # col P
    assert ws.cell(row=4, column=25).value == "water_volume"  # col Y
    assert ws.cell(row=4, column=27).value == "aux1_volume"  # col AA

    # Data starts at row 5 -- water meters first (10), then heat (10), then central (2)
    # First water meter
    assert ws.cell(row=5, column=6).value == "Acqua calda"  # device_detail = col F
    # First heat allocator (after 10 water meters)
    assert ws.cell(row=15, column=6).value == "Contacalorie"
    # Central meters at end (after 10 water + 10 heat = row 25)
    desc_25 = ws.cell(row=25, column=5).value
    assert desc_25 in ("Riscaldamento", "Sanitario")


def test_dati_report_hidden_columns(tmp_path):
    """Irrelevant columns should be hidden."""
    from siemens_converter.parser import parse_fc_report

    fixture = Path(__file__).parent / "fixtures" / "FC_report_TEST_9999_2026-01-15.xls"
    report = parse_fc_report(fixture)
    out = tmp_path / "output.xlsx"
    write_xlsx(report, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Dati Report"]

    # Visible columns: A(1), B(2), E(5), G(7), M(13), P(16), Q(17), R(18), Y(25), Z(26), AA(27), AB(28)
    assert ws.column_dimensions["A"].hidden is False
    assert ws.column_dimensions["E"].hidden is False
    assert ws.column_dimensions["P"].hidden is False
    # Hidden columns
    assert ws.column_dimensions["C"].hidden is True  # device_serial_number
    assert ws.column_dimensions["D"].hidden is True  # name_device
    assert ws.column_dimensions["H"].hidden is True  # wired/wireless


def test_dati_report_formatting(tmp_path):
    """Verify formatting: green highlights, alternating rows, freeze panes."""
    from siemens_converter.parser import parse_fc_report

    fixture = Path(__file__).parent / "fixtures" / "FC_report_TEST_9999_2026-01-15.xls"
    report = parse_fc_report(fixture)
    out = tmp_path / "output.xlsx"
    write_xlsx(report, out)

    wb = openpyxl.load_workbook(out)
    ws = wb["Dati Report"]

    # Header row bold
    assert ws["A1"].font.bold is True
    assert ws.cell(row=4, column=1).font.bold is True

    # Green highlight on water_volume (col Y=25) for first water meter (row 5)
    water_cell = ws.cell(row=5, column=25)
    assert water_cell.fill.fgColor.rgb == "FF99FFCC"

    # Green highlight on heat_energy (col P=16) for first heat allocator (row 15)
    heat_cell = ws.cell(row=15, column=16)
    assert heat_cell.fill.fgColor.rgb == "FF99FFCC"

    # Alternating row fill on even data rows
    alt_cell = ws.cell(row=6, column=1)  # second data row
    assert alt_cell.fill.fgColor.rgb == "FFE8E8E8"

    # Freeze panes
    assert ws.freeze_panes == "A5"

    # Row heights
    assert ws.row_dimensions[5].height == 27


def test_full_pipeline(tmp_path):
    """Parse fixture -> write XLSX -> verify key values."""
    from siemens_converter.parser import parse_fc_report

    fixture = Path(__file__).parent / "fixtures" / "FC_report_TEST_9999_2026-01-15.xls"
    report = parse_fc_report(fixture)
    out = tmp_path / "output.xlsx"
    write_xlsx(report, out)

    wb = openpyxl.load_workbook(out)
    ws = wb.worksheets[0]

    # Central meters populated
    assert ws["C28"].value is not None  # Riscaldamento
    assert ws["C31"].value is not None  # Sanitario
    assert isinstance(ws["C28"].value, int)
    assert isinstance(ws["C31"].value, int)

    # All 10 apartments should have heat readings
    heat_rows = [78, 80, 82, 84, 86, 88, 90, 92, 94, 95]
    for r in heat_rows:
        assert ws.cell(row=r, column=3).value is not None, (
            f"C{r} should have heat value"
        )

    # All 10 apartments should have water readings
    water_rows = [130, 132, 134, 136, 138, 140, 142, 144, 146, 147]
    for r in water_rows:
        assert ws.cell(row=r, column=3).value is not None, (
            f"C{r} should have water value"
        )

    # All 10 apartments should have AFS readings
    afs_rows = [182, 184, 186, 188, 190, 192, 194, 196, 198, 199]
    for r in afs_rows:
        assert ws.cell(row=r, column=3).value is not None, f"C{r} should have AFS value"

    # Dates should be set
    assert ws["D28"].value is not None
    assert ws["C76"].value is not None
