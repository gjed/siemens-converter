import openpyxl
from importlib.resources import files


def _load_template():
    template_path = files("siemens_converter").joinpath("template.xlsx")
    return openpyxl.load_workbook(template_path)


def test_template_has_three_sheets():
    wb = _load_template()
    assert len(wb.sheetnames) == 3


def test_template_ripartizione_sheet_exists():
    wb = _load_template()
    # First sheet name should contain "Ripartizione" or similar
    assert any("Ripartizione" in name or "Riparto" in name for name in wb.sheetnames)


def test_template_central_meter_labels_exist():
    wb = _load_template()
    ws = wb.worksheets[0]
    # Row 27 has the riscaldamento initial reading label
    assert ws["A27"].value is not None
    # Row 30 has the ACS initial reading label
    assert ws["A30"].value is not None


def test_template_data_cells_are_cleared():
    wb = _load_template()
    ws = wb.worksheets[0]
    # These should be None (cleared for writer to populate)
    assert ws["C28"].value is None  # Riscaldamento finale
    assert ws["C31"].value is None  # ACS finale
    assert ws["C78"].value is None  # App 01 heat
    assert ws["C130"].value is None  # App 01 water
    assert ws["C182"].value is None  # App 01 AFS


def test_template_formulas_preserved():
    wb = _load_template()
    ws = wb.worksheets[0]
    # Consumption formula rows should still have formulas
    # D78 = =+C78-B78 (consumption = finale - iniziale)
    d78 = ws["D78"].value
    assert d78 is not None and "=" in str(d78)


def test_template_tabelle_millesimali_structure():
    wb = _load_template()
    ws = wb["Tabelle millesimali"]
    assert ws["A1"].value == "Condomini"
    assert ws["C2"].value == "kWh"


def test_template_tabelle_millesimali_data_cleared():
    wb = _load_template()
    ws = wb["Tabelle millesimali"]
    # Apartment names and data values should be cleared
    for r in range(4, 14):
        assert ws.cell(row=r, column=1).value is None  # names
        assert ws.cell(row=r, column=2).value is None  # subalterno
        assert ws.cell(row=r, column=3).value is None  # energia risc.
        assert ws.cell(row=r, column=5).value is None  # energia ACS
    # But formulas in D and F should still be there
    assert ws["D4"].value is not None
    assert ws["F4"].value is not None


def test_template_tabella_2026_eur_cleared():
    wb = _load_template()
    ws = wb["Tabella_2026"]
    # EUR cost values should be cleared
    for r in range(3, 9):
        assert ws.cell(row=r, column=5).value is None
    # Meter readings should be cleared
    for r in range(20, 23):
        assert ws.cell(row=r, column=6).value is None  # initial
        assert ws.cell(row=r, column=7).value is None  # final
    # Date should be stub
    assert ws["E18"].value is None
    assert ws["F18"].value is None
