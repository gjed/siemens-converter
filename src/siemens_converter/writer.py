"""XLSX workbook generation from parsed FC_report data."""

from __future__ import annotations

import re
import shutil
from datetime import datetime
from importlib.resources import files
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from siemens_converter.models import ParsedReport

# -- Styles matching the original "File con dati necessari" formatting --

_FONT = Font(name="Aptos Narrow", size=10)
_FONT_BOLD = Font(name="Aptos Narrow", size=10, bold=True)
_FONT_HEADER_ANNOTATION = Font(
    name="Aptos Narrow", size=10, bold=True, color="FFFF0000"
)

_FILL_HEADER = PatternFill(
    start_color="FFDBE8F0", end_color="FFDBE8F0", fill_type="solid"
)
_FILL_GREEN = PatternFill(
    start_color="FF99FFCC", end_color="FF99FFCC", fill_type="solid"
)
_FILL_YELLOW = PatternFill(
    start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
)
_FILL_ALT_ROW = PatternFill(
    start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid"
)

_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Row mapping: apartment number -> {heat_row, water_row, afs_row} in Ripartizione sheet
APT_ROWS: dict[int, dict[str, int]] = {
    1: {"heat": 78, "water": 130, "afs": 182},
    2: {"heat": 80, "water": 132, "afs": 184},
    3: {"heat": 82, "water": 134, "afs": 186},
    4: {"heat": 84, "water": 136, "afs": 188},
    5: {"heat": 86, "water": 138, "afs": 190},
    6: {"heat": 88, "water": 140, "afs": 192},
    7: {"heat": 90, "water": 142, "afs": 194},
    8: {"heat": 92, "water": 144, "afs": 196},
    9: {"heat": 94, "water": 146, "afs": 198},
    10: {"heat": 95, "water": 147, "afs": 199},
}


def write_xlsx(report: ParsedReport, output_path: Path) -> None:
    """Write parsed report data into a copy of the template XLSX."""
    template = files("siemens_converter").joinpath("template.xlsx")
    shutil.copy(str(template), str(output_path))

    wb = openpyxl.load_workbook(output_path)
    ws = wb.worksheets[0]  # Ripartizione sheet

    reading_date = _parse_reading_date(report)

    # Central meter readings
    for meter in report.central_meters:
        if meter.description == "Riscaldamento":
            ws["C28"] = meter.heat_energy_kwh
            ws["D28"] = reading_date
        elif meter.description == "Sanitario":
            ws["C31"] = meter.heat_energy_kwh
            ws["D31"] = reading_date

    # Section date headers
    ws["C76"] = reading_date
    ws["C126"] = reading_date
    ws["C178"] = reading_date

    # Per-apartment heat allocator readings
    for ha in report.heat_allocators:
        rows = APT_ROWS.get(ha.apartment_number)
        if rows is None:
            continue
        ws.cell(row=rows["heat"], column=3, value=ha.heat_energy_kwh)
        ws.cell(row=rows["afs"], column=3, value=ha.aux1_volume_m3)

    # Per-apartment water meter readings
    for wm in report.water_meters:
        rows = APT_ROWS.get(wm.apartment_number)
        if rows is None:
            continue
        ws.cell(row=rows["water"], column=3, value=wm.water_volume_m3)

    # Populate apartment names throughout the workbook
    _write_apartment_names(wb, report)

    # Add "Inquilini" sheet (empty tenant table for manual input)
    _write_inquilini_sheet(wb, report)

    # Add "Dati Report" sheet with raw FC_report data
    _write_dati_report_sheet(wb, report)

    wb.save(output_path)


# Millesimali sheet: apartment number -> row (rows 4-13)
_MILL_ROWS: dict[int, int] = {i: i + 3 for i in range(1, 11)}


# Summary rows in Ripartizione: apartment N -> row N+2 (row 3=apt 1, row 4=apt 2...)
_SUMMARY_ROWS: dict[int, int] = {i: i + 2 for i in range(1, 11)}


def _write_apartment_names(wb: openpyxl.Workbook, report: ParsedReport) -> None:
    """Write FC_report apartment descriptions to summary rows and millesimali."""
    apt_names: dict[int, str] = {}
    for ha in report.heat_allocators:
        apt_names[ha.apartment_number] = ha.description

    # Ripartizione: only summary rows 3-12 (other rows use formulas referencing these)
    ws = wb.worksheets[0]
    for apt_num, name in apt_names.items():
        row = _SUMMARY_ROWS.get(apt_num)
        if row is not None:
            ws.cell(row=row, column=1, value=name)

    # Tabelle millesimali
    ws_mill = wb["Tabelle millesimali"]
    for apt_num, name in apt_names.items():
        row = _MILL_ROWS.get(apt_num)
        if row is not None:
            ws_mill.cell(row=row, column=1, value=name)


def _write_inquilini_sheet(wb: openpyxl.Workbook, report: ParsedReport) -> None:
    """Create an 'Inquilini' sheet with apartment list and empty tenant column."""
    ws = wb.create_sheet("Inquilini")  # Appended after other sheets

    # Header row
    ws.cell(row=1, column=1, value="Appartamento").font = _FONT_BOLD
    ws.cell(row=1, column=1).fill = _FILL_HEADER
    ws.cell(row=1, column=1).alignment = _ALIGN_LEFT
    ws.cell(row=1, column=1).border = _THIN_BORDER

    ws.cell(row=1, column=2, value="Inquilino").font = _FONT_BOLD
    ws.cell(row=1, column=2).fill = _FILL_HEADER
    ws.cell(row=1, column=2).alignment = _ALIGN_LEFT
    ws.cell(row=1, column=2).border = _THIN_BORDER

    # One row per apartment from heat allocators (sorted by apt number)
    for i, ha in enumerate(report.heat_allocators):
        row = i + 2
        c = ws.cell(row=row, column=1, value=ha.description)
        c.font = _FONT
        c.alignment = _ALIGN_LEFT
        c.border = _THIN_BORDER

        # Empty tenant cell for manual input
        c2 = ws.cell(row=row, column=2, value="")
        c2.font = _FONT
        c2.alignment = _ALIGN_LEFT
        c2.border = _THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42

    # Row heights
    ws.row_dimensions[1].height = 30
    for i in range(len(report.heat_allocators)):
        ws.row_dimensions[i + 2].height = 25


# Columns visible in the original "File con dati necessari" (0-indexed)
_VISIBLE_COLS = {0, 1, 4, 6, 12, 15, 16, 17, 24, 25, 26, 27}

# Columns that get green highlight (key data)
_GREEN_COLS = {15, 24, 26}  # heat_energy, water_volume, aux1_volume

# Column that gets yellow highlight (device_description)
_YELLOW_COL = 4

# Minimum column widths for visible columns (0-indexed col -> width)
_MIN_COL_WIDTHS: dict[int, float] = {
    0: 11,
    1: 11,
    4: 40,
    6: 32,
    12: 16,
    15: 14,
    16: 8,
    17: 14,
    24: 14,
    25: 8,
    26: 14,
    27: 8,
}

# Row 1 group headers above data columns (0-indexed col -> label)
_FILL_BRIGHT_GREEN = PatternFill(
    start_color="FF66FF33",
    end_color="FF66FF33",
    fill_type="solid",
)
_ROW1_GROUP_HEADERS: dict[int, tuple[str, PatternFill]] = {
    15: ("Riparto", _FILL_BRIGHT_GREEN),
    24: ("Riparto", _FILL_BRIGHT_GREEN),
    26: ("Conti separati a parte", _FILL_BRIGHT_GREEN),
}

# Row 3 sub-labels for data column groups (0-indexed col -> label, fill)
_ROW3_ANNOTATIONS: dict[int, tuple[str, PatternFill]] = {
    15: ("energia termica", _FILL_GREEN),
    16: ("unita\u0300 energia calore", _FILL_GREEN),
    17: ("NO energia fresca Split_pers", _FILL_GREEN),
    24: ("volume acqua sanitaria", _FILL_GREEN),
    25: ("unita\u0300 volume acqua", _FILL_GREEN),
    26: ("volume acqua fredda", _FILL_YELLOW),
}

# Metadata: placed in the FC_report's own column positions (0-indexed)
_META_FIELDS: list[tuple[int, str]] = [
    (0, "Nome File"),
    (1, "Data Report"),
    (2, "Ora Report"),
    (3, "Riferimento Impianto"),
    (4, "Versione firmware"),
    (5, "Totale dispositivi cablati"),
    (7, "Numero di serie"),
]


def _write_dati_report_sheet(wb: openpyxl.Workbook, report: ParsedReport) -> None:
    """Add a 'Dati Report' sheet — full FC_report table with irrelevant columns hidden."""
    ws = wb.create_sheet("Dati Report")

    headers = report.column_headers or []
    raw_rows = report.raw_device_rows or []
    num_cols = max(len(headers), 38)

    # -- Row 1: metadata labels + group headers over data columns --
    # Metadata labels in FC_report column positions
    meta_values = [
        report.header.filename,
        report.header.report_date,
        report.header.report_time,
        report.header.reference,
        report.header.firmware,
        str(report.header.total_wired),
        report.header.serial,
    ]
    for i, (col_idx, label) in enumerate(_META_FIELDS):
        c = ws.cell(row=1, column=col_idx + 1, value=label)
        c.font = _FONT_BOLD
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _THIN_BORDER
    # Group headers over the data columns (same row 1)
    for col_idx, (label, fill) in _ROW1_GROUP_HEADERS.items():
        c = ws.cell(row=1, column=col_idx + 1, value=label)
        c.font = _FONT_BOLD
        c.fill = fill
        c.alignment = _ALIGN_CENTER
        c.border = _THIN_BORDER
    ws.row_dimensions[1].height = 35

    # -- Row 2: metadata values --
    for i, (col_idx, _) in enumerate(_META_FIELDS):
        c = ws.cell(row=2, column=col_idx + 1, value=meta_values[i])
        c.font = _FONT
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _THIN_BORDER
    ws.row_dimensions[2].height = 40

    # -- Row 3: sub-labels for data column groups --
    for col_idx, (label, fill) in _ROW3_ANNOTATIONS.items():
        c = ws.cell(row=3, column=col_idx + 1, value=label)
        c.font = _FONT_HEADER_ANNOTATION
        c.fill = fill
        c.alignment = _ALIGN_CENTER
        c.border = _THIN_BORDER
    ws.row_dimensions[3].height = 37

    # -- Row 4: column headers (all 38 from FC_report) --
    for col_idx, h in enumerate(headers):
        c = ws.cell(row=4, column=col_idx + 1, value=h)
        c.font = _FONT_BOLD
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_LEFT
        c.border = _THIN_BORDER
        if col_idx == _YELLOW_COL:
            c.fill = _FILL_YELLOW
    ws.row_dimensions[4].height = 35

    # -- Track max content width per column for auto-fit --
    col_max_width: dict[int, float] = {}
    for col_idx, h in enumerate(headers):
        col_max_width[col_idx] = len(str(h)) * 1.2 if h else 8

    # -- Data rows (all raw device rows with all columns) --
    for row_offset, raw_row in enumerate(raw_rows):
        xl_row = 5 + row_offset
        is_alt = row_offset % 2 == 1

        for col_idx in range(min(len(raw_row), num_cols)):
            val = raw_row[col_idx] if col_idx < len(raw_row) else ""
            parsed_val = _try_parse_number(val)

            c = ws.cell(row=xl_row, column=col_idx + 1, value=parsed_val)
            c.border = _THIN_BORDER

            # Track content width
            content_len = len(str(parsed_val)) * 1.2 + 2 if parsed_val else 0
            col_max_width[col_idx] = max(
                col_max_width.get(col_idx, 0),
                content_len,
            )

            # Styling
            if col_idx in _GREEN_COLS and parsed_val not in (None, "", 0, 0.0):
                c.font = _FONT_BOLD
                c.fill = _FILL_GREEN
            elif col_idx == _YELLOW_COL:
                c.font = _FONT_BOLD
            elif is_alt:
                c.font = _FONT
                c.fill = _FILL_ALT_ROW
            else:
                c.font = _FONT

            # Alignment
            if isinstance(parsed_val, (int, float)):
                c.alignment = _ALIGN_RIGHT
                if isinstance(parsed_val, float):
                    c.number_format = "0.000"
            else:
                c.alignment = _ALIGN_LEFT

        ws.row_dimensions[xl_row].height = 27

    # -- Column widths: use max(min_width, content_width) and visibility --
    for col_idx in range(num_cols):
        letter = get_column_letter(col_idx + 1)
        min_w = _MIN_COL_WIDTHS.get(col_idx, 12)
        content_w = col_max_width.get(col_idx, 8)
        ws.column_dimensions[letter].width = max(min_w, content_w)
        if col_idx not in _VISIBLE_COLS:
            ws.column_dimensions[letter].hidden = True

    # -- Freeze panes below headers --
    ws.freeze_panes = "A5"

    # -- Auto-filter on the data table --
    last_row = 4 + len(raw_rows)
    last_col_letter = get_column_letter(num_cols)
    if raw_rows:
        ws.auto_filter.ref = f"A4:{last_col_letter}{last_row}"


def _try_parse_number(val: str) -> str | int | float:
    """Try to parse a string as int or float (Italian comma decimals)."""
    if not isinstance(val, str):
        return val
    val = val.strip()
    if not val:
        return ""
    # Try int first
    try:
        return int(val)
    except ValueError:
        pass
    # Try float with comma decimal
    try:
        return float(val.replace(",", "."))
    except ValueError:
        return val


def _parse_reading_date(report: ParsedReport) -> datetime | None:
    """Extract a datetime from the first available readout_date string."""
    for meter in report.central_meters:
        if meter.readout_date:
            try:
                return datetime.strptime(meter.readout_date, "%Y/%m/%d")
            except ValueError:
                pass
    return None
